#!/usr/bin/env python3
import json
import sys
from html import escape
from pathlib import Path
from typing import Any


def _normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).replace("\r", "\n").strip()


def _trim_empty_edges(rows: list[list[str]]) -> list[list[str]]:
    while rows and not any(cell.strip() for cell in rows[0]):
        rows.pop(0)
    while rows and not any(cell.strip() for cell in rows[-1]):
        rows.pop()
    if not rows:
        return []

    max_cols = max(len(row) for row in rows)
    normalized = [row + [""] * (max_cols - len(row)) for row in rows]
    keep_cols = [idx for idx in range(max_cols) if any(row[idx].strip() for row in normalized)]
    return [[row[idx] for idx in keep_cols] for row in normalized]


def _rows_to_tsv(rows: list[list[str]]) -> str:
    return "\n".join("\t".join(cell.replace("\n", " / ") for cell in row) for row in rows)


def _rows_to_html(rows: list[list[str]], sheet_name: str) -> str:
    lines = [f'<table sheet="{escape(sheet_name, quote=True)}">']
    for row in rows:
        lines.append("  <tr>")
        for cell in row:
            lines.append(f"    <td>{escape(cell).replace(chr(10), '<br>')}</td>")
        lines.append("  </tr>")
    lines.append("</table>")
    return "\n".join(lines)


def _xlsx_sheets(path: Path) -> list[tuple[str, list[list[str]]]]:
    from openpyxl import load_workbook  # type: ignore

    workbook = load_workbook(path, data_only=True, read_only=True)
    sheets: list[tuple[str, list[list[str]]]] = []
    for sheet in workbook.worksheets:
        rows = [[_normalize_cell(cell) for cell in row] for row in sheet.iter_rows(values_only=True)]
        sheets.append((sheet.title, _trim_empty_edges(rows)))
    return sheets


def _xls_sheets(path: Path) -> list[tuple[str, list[list[str]]]]:
    try:
        import xlrd  # type: ignore
    except Exception as exc:
        raise RuntimeError("解析 .xls 需要安装 xlrd：python3 -m pip install xlrd") from exc

    workbook = xlrd.open_workbook(str(path))
    sheets: list[tuple[str, list[list[str]]]] = []
    for sheet in workbook.sheets():
        rows = [[_normalize_cell(sheet.cell_value(r, c)) for c in range(sheet.ncols)] for r in range(sheet.nrows)]
        sheets.append((sheet.name, _trim_empty_edges(rows)))
    return sheets


def _extract(path: Path) -> dict[str, Any]:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        sheets = _xlsx_sheets(path)
    elif suffix == ".xls":
        sheets = _xls_sheets(path)
    else:
        raise RuntimeError(f"不支持的 Excel 扩展名：{suffix}")

    parts: list[str] = []
    table_count = 0
    row_count = 0
    for sheet_name, rows in sheets:
        if not rows:
            continue
        table_count += 1
        row_count += len(rows)
        parts.append(f"--- 工作表：{sheet_name} ---")
        parts.append(f"[表格 parser=excel sheet={sheet_name}]")
        parts.append(_rows_to_html(rows, sheet_name))
        parts.append("[TSV]")
        parts.append(_rows_to_tsv(rows))

    text = "\n".join(part for part in parts if part).strip()
    return {
        "text": text,
        "sheet_count": len(sheets),
        "table_count": table_count,
        "row_count": row_count,
        "errors": [] if text else ["Excel 未解析到非空单元格"],
    }


def main() -> None:
    if len(sys.argv) != 2:
        raise SystemExit("用法：excel-extract.py <Excel路径>")
    result = _extract(Path(sys.argv[1]))
    print(json.dumps(result, ensure_ascii=False))


if __name__ == "__main__":
    main()
