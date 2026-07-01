from __future__ import annotations

import os
from pathlib import Path
from typing import Any

EXCEL_ATTACHMENT_ROW_THRESHOLD = 5
PDF_OCR_MAX_PAGES = 10
OCR_IMAGE_EXTENSIONS = (".jpg", ".jpeg", ".png", ".bmp", ".gif", ".tif", ".tiff", ".webp")


def _normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).replace("\r", "\n").strip()


def _trim_empty_edges(rows: list[list[str]]) -> list[list[str]]:
    while rows and not any(cell.strip() for cell in rows[0]):
        rows.pop(0)
    while rows and not any(cell.strip() for cell in rows[-1]):
        rows.pop()
    if not rows:
        return []
    max_cols = max(len(row) for row in rows)
    normalized = [row + [""] * (max_cols - len(row)) for row in rows]
    keep_cols = [idx for idx in range(max_cols) if any(row[idx].strip() for row in normalized)]
    return [[row[idx] for idx in keep_cols] for row in normalized]


def _rows_to_tsv(rows: list[list[str]]) -> str:
    return "\n".join("\t".join(cell.replace("\n", " / ") for cell in row) for row in rows)


def _read_excel_sheets(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    sheets: list[dict[str, Any]] = []
    try:
        if suffix == ".xlsx":
            from openpyxl import load_workbook

            workbook = load_workbook(path, data_only=True, read_only=True)
            for sheet in workbook.worksheets:
                rows = [[_normalize_cell(cell) for cell in row] for row in sheet.iter_rows(values_only=True)]
                trimmed_rows = _trim_empty_edges(rows)
                sheets.append({"name": sheet.title, "rows": trimmed_rows, "rowCount": len(trimmed_rows)})
        elif suffix == ".xls":
            import xlrd

            workbook = xlrd.open_workbook(str(path))
            for sheet in workbook.sheets():
                rows = [[_normalize_cell(sheet.cell_value(r, c)) for c in range(sheet.ncols)] for r in range(sheet.nrows)]
                trimmed_rows = _trim_empty_edges(rows)
                sheets.append({"name": sheet.name, "rows": trimmed_rows, "rowCount": len(trimmed_rows)})
        else:
            raise ValueError(f"不支持的 Excel 扩展名：{suffix}")
    except ValueError:
        raise
    except Exception as exc:
        raise ValueError("Excel 报价单格式不正确或文件已损坏，请重新上传") from exc
    return sheets


def excel_attachment_mode(sheets: list[dict[str, Any]], row_threshold: int = EXCEL_ATTACHMENT_ROW_THRESHOLD) -> dict[str, Any]:
    non_empty_sheets = [sheet for sheet in sheets if sheet.get("rows")]
    sheet_count = len(non_empty_sheets)
    row_counts = [int(sheet.get("rowCount") or len(sheet.get("rows") or [])) for sheet in non_empty_sheets]
    total_rows = sum(row_counts)
    max_rows = max(row_counts, default=0)
    reasons: list[str] = []
    if sheet_count > 1:
        reasons.append("multiple_sheets")
    if max_rows > row_threshold:
        reasons.append("sheet_rows_over_threshold")
    if total_rows > row_threshold:
        reasons.append("total_rows_over_threshold")
    return {
        "enabled": bool(reasons),
        "reasons": reasons,
        "rowThreshold": row_threshold,
        "sheetCount": sheet_count,
        "rowCount": total_rows,
        "maxSheetRowCount": max_rows,
    }


def extract_excel_payload(path: Path) -> dict[str, Any]:
    sheets = _read_excel_sheets(path)
    parts: list[str] = []
    for sheet in sheets:
        rows = sheet.get("rows") or []
        if not rows:
            continue
        sheet_name = str(sheet.get("name") or "Sheet")
        parts.append(f"--- 工作表：{sheet_name} ---")
        parts.append(f"[表格 parser=excel sheet={sheet_name} format=tsv]")
        parts.append(_rows_to_tsv(rows))
    text = "\n".join(parts).strip()
    if not text:
        raise ValueError("Excel 未解析到非空单元格")
    return {
        "quoteText": text,
        "sheets": sheets,
        "attachmentMode": excel_attachment_mode(sheets),
    }


def _response_to_map(response: Any) -> dict[str, Any]:
    if hasattr(response, "to_map"):
        value = response.to_map()
        return value if isinstance(value, dict) else {}
    if isinstance(response, dict):
        return response
    return {}


def _flatten_text(value: Any) -> str:
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, dict):
        return "\n".join(part for part in (_flatten_text(item) for item in value.values()) if part)
    if isinstance(value, list):
        return "\n".join(part for part in (_flatten_text(item) for item in value) if part)
    return ""


def extract_excel_text(path: Path) -> str:
    return str(extract_excel_payload(path)["quoteText"])


def _env_flag(name: str, default: bool = False) -> bool:
    value = (os.getenv(name) or "").strip().lower()
    if not value:
        return default
    return value in {"1", "true", "yes", "on"}


def _ocr_scene() -> str:
    scene = (os.getenv("ALIYUN_OCR_SCENE") or "Advanced").strip()
    return scene if scene in {"Advanced", "General"} else "Advanced"


def _is_image_file(lower_name: str, mime_type: str) -> bool:
    return lower_name.endswith(OCR_IMAGE_EXTENSIONS) or mime_type.startswith("image/")


def _normalize_ocr_cell_content(value: Any) -> str:
    text = _normalize_cell(value)
    if len(text) >= 2 and text[0] == text[-1] == '"':
        return text[1:-1].strip()
    return text


def _ocr_table_rows(table_detail: dict[str, Any]) -> list[list[str]]:
    cells = table_detail.get("CellDetails") or []
    if not cells:
        return []
    max_row = max(int(cell.get("RowEnd") or cell.get("RowStart") or 0) for cell in cells)
    max_col = max(int(cell.get("ColumnEnd") or cell.get("ColumnStart") or 0) for cell in cells)
    grid = [["" for _ in range(max_col + 1)] for _ in range(max_row + 1)]
    for cell in cells:
        row_start = int(cell.get("RowStart") or 0)
        row_end = int(cell.get("RowEnd") or row_start)
        col_start = int(cell.get("ColumnStart") or 0)
        col_end = int(cell.get("ColumnEnd") or col_start)
        content = _normalize_ocr_cell_content(cell.get("CellContent", ""))
        for row_index in range(row_start, row_end + 1):
            for col_index in range(col_start, col_end + 1):
                if not grid[row_index][col_index]:
                    grid[row_index][col_index] = content
                elif content and content not in grid[row_index][col_index]:
                    grid[row_index][col_index] = f"{grid[row_index][col_index]} {content}".strip()
    return _trim_empty_edges(grid)


def _ocr_response_data(response: Any) -> dict[str, Any]:
    raw = _response_to_map(response)
    data = raw.get("body") if isinstance(raw.get("body"), dict) else raw
    content = data.get("Data") or data.get("data") or data.get("content") if isinstance(data, dict) else None
    return content if isinstance(content, dict) else {}


def _ocr_data_to_page_text(data: dict[str, Any], page_no: int, scene: str) -> str:
    parser_name = f"ocr-{scene.lower()}"
    parts = [f"--- 第 {page_no} 页 ---"]
    table_info = data.get("TableInfo") or {}
    for table_no, table in enumerate(table_info.get("TableDetails") or [], start=1):
        rows = _ocr_table_rows(table if isinstance(table, dict) else {})
        if rows:
            parts.append(f"[表格 parser={parser_name} page={page_no} index={table_no} format=tsv]")
            parts.append(_rows_to_tsv(rows))
    content = _flatten_text(data.get("Content"))
    if content:
        parts.append(f"[第 {page_no} 页文字]")
        parts.append(content)
    return "\n".join(parts).strip()


def _build_ocr_client() -> Any:
    try:
        from alibabacloud_ocr_api20210707.client import Client as OcrClient
        from alibabacloud_tea_openapi import models as open_api_models
    except ImportError as exc:
        raise ValueError("未安装阿里云 OCR SDK，无法识别图片报价单") from exc

    access_key_id = (os.getenv("ALIYUN_ACCESS_KEY_ID") or "").strip()
    access_key_secret = (os.getenv("ALIYUN_ACCESS_KEY_SECRET") or "").strip()
    endpoint = (os.getenv("ALIYUN_OCR_ENDPOINT") or "ocr-api.cn-hangzhou.aliyuncs.com").strip()
    region_id = (os.getenv("ALIYUN_OCR_REGION_ID") or "cn-hangzhou").strip()
    if not access_key_id or not access_key_secret:
        raise ValueError("未配置阿里云 OCR 访问凭证")
    config = open_api_models.Config(
        access_key_id=access_key_id,
        access_key_secret=access_key_secret,
        endpoint=endpoint,
        region_id=region_id,
    )
    return OcrClient(config)


def _recognize_all_text(path: Path, page_no: int | None = None) -> str:
    try:
        from alibabacloud_ocr_api20210707 import models as ocr_models
        from alibabacloud_tea_util import models as util_models
    except ImportError as exc:
        raise ValueError("未安装阿里云 OCR SDK，无法识别图片报价单") from exc

    scene = _ocr_scene()
    output_table = _env_flag("ALIYUN_OCR_OUTPUT_TABLE", default=True)
    lineless_table = _env_flag("ALIYUN_OCR_LINELESS_TABLE", default=False)
    client = _build_ocr_client()
    advanced_config = None
    if scene == "Advanced":
        advanced_config = ocr_models.RecognizeAllTextRequestAdvancedConfig(
            output_table=output_table,
            is_line_less_table=lineless_table,
        )
    try:
        with path.open("rb") as body:
            request_kwargs: dict[str, Any] = {"type": scene, "body": body}
            if page_no is not None:
                request_kwargs["page_no"] = page_no
            if advanced_config is not None:
                request_kwargs["advanced_config"] = advanced_config
            request = ocr_models.RecognizeAllTextRequest(**request_kwargs)
            response = client.recognize_all_text_with_options(request, util_models.RuntimeOptions())
    except Exception as exc:
        raise ValueError(f"图片 OCR 识别失败：{exc}") from exc

    data = _ocr_response_data(response)
    if not data:
        return ""
    return _ocr_data_to_page_text(data, page_no or 1, scene)


def _extract_pdf_with_pdfplumber(path: Path) -> str:
    import pdfplumber

    table_settings = {
        "vertical_strategy": "lines",
        "horizontal_strategy": "lines",
        "snap_tolerance": 3,
        "join_tolerance": 3,
        "intersection_tolerance": 5,
        "text_x_tolerance": 2,
        "text_y_tolerance": 3,
    }
    text_table_settings = {
        "vertical_strategy": "text",
        "horizontal_strategy": "text",
        "snap_tolerance": 3,
        "join_tolerance": 3,
        "intersection_tolerance": 5,
        "min_words_vertical": 2,
        "min_words_horizontal": 1,
        "text_x_tolerance": 2,
        "text_y_tolerance": 3,
    }
    parts: list[str] = []
    with pdfplumber.open(path) as pdf:
        for page_no, page in enumerate(pdf.pages, start=1):
            parts.append(f"--- 第 {page_no} 页 ---")
            page_tables = page.find_tables(table_settings=table_settings)
            parser_name = "pdfplumber-lines"
            if not page_tables:
                page_tables = page.find_tables(table_settings=text_table_settings)
                parser_name = "pdfplumber-text"
            for table_no, table in enumerate(page_tables, start=1):
                rows = [[_normalize_cell(cell) for cell in row] for row in (table.extract() or []) if row]
                if rows:
                    parts.append(f"[表格 parser={parser_name} page={page_no} index={table_no} format=tsv]")
                    parts.append(_rows_to_tsv(rows))
            text = (page.extract_text(x_tolerance=2, y_tolerance=3, layout=True) or "").strip()
            if text:
                parts.append(f"[第 {page_no} 页文字]")
                parts.append(text)
    result = "\n".join(parts).strip()
    if not result:
        raise ValueError("PDF 未解析到文本")
    return result


def _extract_pdf_with_ocr(path: Path) -> str:
    import pdfplumber

    with pdfplumber.open(path) as pdf:
        page_count = min(len(pdf.pages), PDF_OCR_MAX_PAGES)
    parts: list[str] = []
    for page_no in range(1, page_count + 1):
        page_text = _recognize_all_text(path, page_no=page_no)
        if page_text:
            parts.append(page_text)
    result = "\n".join(parts).strip()
    if not result:
        raise ValueError("PDF OCR 未识别到有效文本")
    return result


def _pdfplumber_has_content(text: str) -> bool:
    return "[表格 parser=" in text or "[第 " in text and "页文字]" in text


def _extract_pdf_content(path: Path) -> tuple[str, bool]:
    try:
        text = _extract_pdf_with_pdfplumber(path)
    except ValueError:
        text = ""
    if text.strip() and _pdfplumber_has_content(text):
        return text, False
    return _extract_pdf_with_ocr(path), True


def extract_pdf_text(path: Path) -> str:
    text, _ocr_used = _extract_pdf_content(path)
    return text


def extract_image_text(path: Path) -> str:
    text = _recognize_all_text(path)
    if not text:
        raise ValueError("图片 OCR 未识别到有效文本")
    return text


def parser_metadata_for_file(path: Path, mime_type: str = "") -> dict[str, Any]:
    lower = path.name.lower()
    if lower.endswith((".xlsx", ".xls")) or "spreadsheet" in mime_type or "ms-excel" in mime_type:
        return {"type": "excel", "ocrUsed": False}
    if lower.endswith(".pdf") or mime_type == "application/pdf":
        return {"type": "pdf", "ocrUsed": False}
    if _is_image_file(lower, mime_type):
        return {"type": "image", "ocrUsed": True}
    if mime_type.startswith("text/") or lower.endswith(".txt"):
        return {"type": "text", "ocrUsed": False}
    return {"type": "unknown", "ocrUsed": False}


def extract_quote_content(path: Path, mime_type: str = "") -> tuple[str, dict[str, Any]]:
    lower = path.name.lower()
    if mime_type.startswith("text/") or lower.endswith(".txt"):
        text = path.read_text(encoding="utf-8").strip()
        if not text:
            raise ValueError("文本报价单内容为空")
        return text, {}
    if lower.endswith((".xlsx", ".xls")) or "spreadsheet" in mime_type or "ms-excel" in mime_type:
        return extract_excel_text(path), {}
    if lower.endswith(".pdf") or mime_type == "application/pdf":
        text, ocr_used = _extract_pdf_content(path)
        return text, {"ocrUsed": True} if ocr_used else {}
    if _is_image_file(lower, mime_type):
        return extract_image_text(path), {"ocrUsed": True}
    raise ValueError("当前版本支持 PDF、Excel 或图片报价单")


def extract_text_from_file(path: Path, mime_type: str = "") -> str:
    quote_text, _parser_patch = extract_quote_content(path, mime_type)
    return quote_text
