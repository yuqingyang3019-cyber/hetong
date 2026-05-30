from __future__ import annotations

import base64
import hashlib
import hmac
import json
import os
import re
import time
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Any
from urllib.parse import quote, urlencode

import requests
from openpyxl import Workbook, load_workbook


DATA_CENTER_URL = "https://apigateway.yonyoucloud.com/open-auth/dataCenter/getGatewayAddress"
DEFAULT_YONBIP_GATEWAY_URL = "https://c3.yonyoucloud.com/iuap-api-gateway"
DEFAULT_YONBIP_TOKEN_URL = "https://c3.yonyoucloud.com/iuap-api-gateway"
TOKEN_PATH = "/open-auth/selfAppAuth/getAccessToken"
VENDOR_QUERY_PATH = "/yonbip/digitalModel/vendor/queryByPage"
DEFAULT_PAGE_SIZE = 500
SUPPLIER_CACHE_FILE_NAME = "supplier-cache.xlsx"
EXPLICIT_VENDOR_DATA_FIELDS = ",".join([
    "id",
    "code",
    "name",
    "creditcode",
    "address",
    "contactphone",
    "vendorphone",
    "vendorfax",
    "vendoraddress",
    "orgId",
    "org",
    "accessstatus",
    "freezestatus",
    "pubts",
])
VENDOR_COLUMNS = [
    ("id", "供应商ID"),
    ("code", "供应商编码"),
    ("name", "供应商名称"),
    ("creditcode", "统一社会信用代码"),
    ("address", "地址"),
    ("contactphone", "电话"),
    ("openaccountbankName", "开户行"),
    ("bankAccount", "银行账号"),
    ("bankAccountName", "户名"),
    ("vendorFax", "传真"),
    ("org", "组织ID"),
    ("orgName", "组织名称"),
    ("accessstatus", "准入状态"),
    ("freezestatus", "冻结状态"),
    ("pubts", "更新时间"),
]
VENDOR_COLUMN_LABEL_TO_KEY = {label: key for key, label in VENDOR_COLUMNS}
SUPPLIER_FIELD_MAP = {
    "supplierAddress": "address",
    "supplierBank": "openaccountbankName",
    "supplierAccount": "bankAccount",
    "supplierTaxNo": "creditcode",
    "supplierPhone": "contactphone",
    "supplierFax": "vendorFax",
}


def require_env(name: str) -> str:
    value = (os.getenv(name) or "").strip()
    if not value:
        raise RuntimeError(f"缺少配置：{name}")
    return value


def optional_env(name: str) -> str:
    return (os.getenv(name) or "").strip()


def env_int(name: str, default: int) -> int:
    raw = optional_env(name)
    if not raw:
        return default
    try:
        value = int(raw)
    except ValueError as exc:
        raise RuntimeError(f"环境变量 {name} 必须是整数") from exc
    if value <= 0:
        raise RuntimeError(f"环境变量 {name} 必须大于 0")
    return value


def as_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, dict):
        for key in ("zh_CN", "name", "value"):
            nested = value.get(key)
            if nested is not None and str(nested).strip():
                return str(nested).strip()
        for nested in value.values():
            text = as_text(nested)
            if text:
                return text
        return ""
    if isinstance(value, list):
        return " ".join(text for text in (as_text(item) for item in value) if text).strip()
    return str(value).strip()


def is_blank(value: Any) -> bool:
    return not as_text(value)


def normalize_supplier_name(value: Any) -> str:
    text = as_text(value).lower()
    text = re.sub(r"\s+", "", text)
    return re.sub(r"[（）()【】\\[\\]\"'“”‘’.,，。;；:：、]", "", text)


def api_get(url: str, params: dict[str, Any] | None = None) -> dict[str, Any]:
    response = requests.get(url, params=params, timeout=30)
    response.raise_for_status()
    body = response.json()
    if not isinstance(body, dict):
        raise RuntimeError(f"接口返回不是 JSON 对象：{url}")
    return body


def api_post(url: str, params: dict[str, Any] | None = None, payload: dict[str, Any] | None = None) -> dict[str, Any]:
    response = requests.post(url, params=params, json=payload or {}, timeout=60)
    response.raise_for_status()
    body = response.json()
    if not isinstance(body, dict):
        raise RuntimeError(f"接口返回不是 JSON 对象：{url}")
    return body


def success_code(body: dict[str, Any], expected: str) -> bool:
    return str(body.get("code") or "").strip() == expected


def get_gateway_address(tenant_id: str) -> tuple[str, str]:
    body = api_get(DATA_CENTER_URL, {"tenantId": tenant_id})
    if not success_code(body, "00000"):
        raise RuntimeError(f"获取用友数据中心域名失败：{body.get('message') or body}")
    data = body.get("data") if isinstance(body.get("data"), dict) else {}
    gateway_url = str(data.get("gatewayUrl") or "").rstrip("/")
    token_url = str(data.get("tokenUrl") or "").rstrip("/")
    if not gateway_url or not token_url:
        raise RuntimeError("用友数据中心域名返回缺少 gatewayUrl 或 tokenUrl")
    return gateway_url, token_url


def resolve_endpoints() -> tuple[str, str]:
    gateway_url = optional_env("YONBIP_GATEWAY_URL").rstrip("/") or DEFAULT_YONBIP_GATEWAY_URL
    token_url = optional_env("YONBIP_TOKEN_URL").rstrip("/") or DEFAULT_YONBIP_TOKEN_URL
    if gateway_url and token_url:
        return gateway_url, token_url

    tenant_id = optional_env("YONBIP_TENANT_ID")
    if tenant_id:
        return get_gateway_address(tenant_id)
    if gateway_url:
        return gateway_url, gateway_url
    raise RuntimeError("缺少配置：YONBIP_GATEWAY_URL；或配置 YONBIP_TENANT_ID 自动解析动态域名")


def sign_token_request(app_key: str, app_secret: str, timestamp: int) -> str:
    plain = f"appKey{app_key}timestamp{timestamp}"
    digest = hmac.new(app_secret.encode("utf-8"), plain.encode("utf-8"), hashlib.sha256).digest()
    return quote(base64.b64encode(digest).decode("ascii"), safe="")


def get_access_token(token_url: str) -> tuple[str, int]:
    app_key = require_env("YONBIP_APP_KEY")
    app_secret = require_env("YONBIP_APP_SECRET")
    timestamp = int(time.time() * 1000)
    signature = sign_token_request(app_key, app_secret, timestamp)
    url = f"{token_url}{TOKEN_PATH}?{urlencode({'appKey': app_key, 'timestamp': timestamp})}&signature={signature}"
    body = api_get(url)
    if not success_code(body, "00000"):
        raise RuntimeError(f"获取用友 access_token 失败：{body.get('message') or body}")
    data = body.get("data") if isinstance(body.get("data"), dict) else {}
    token = str(data.get("access_token") or "")
    expire = int(data.get("expire") or 0)
    if not token:
        raise RuntimeError("获取用友 access_token 成功但返回为空")
    return token, expire


def vendor_query_payload(page_index: int, page_size: int) -> dict[str, Any]:
    return {
        "data": EXPLICIT_VENDOR_DATA_FIELDS,
        "page": {
            "pageSize": page_size,
            "pageIndex": page_index,
        },
        "queryOrders": [
            {
                "field": "code",
                "order": "asc",
            }
        ],
        "partParam": {
            "vendorbanks": {
                "data": "*,openaccountbank.name",
            }
        },
    }


def query_vendor_page(gateway_url: str, access_token: str, page_index: int, page_size: int) -> dict[str, Any]:
    body = api_post(
        f"{gateway_url}{VENDOR_QUERY_PATH}",
        {"access_token": access_token},
        vendor_query_payload(page_index, page_size),
    )
    if not success_code(body, "200"):
        raise RuntimeError(f"用友供应商分页查询失败：{body.get('message') or body}")
    data = body.get("data") if isinstance(body.get("data"), dict) else {}
    records = data.get("recordList") if isinstance(data.get("recordList"), list) else []
    return {
        "recordCount": int(data.get("recordCount") or 0),
        "pageIndex": int(data.get("pageIndex") or page_index),
        "pageSize": int(data.get("pageSize") or page_size),
        "pageCount": int(data.get("pageCount") or 0),
        "records": [record for record in records if isinstance(record, dict)],
    }


def vendor_is_available(record: dict[str, Any]) -> bool:
    freeze_status = record.get("freezestatus")
    is_frozen = freeze_status is True or str(freeze_status) == "1"
    is_stopped = record.get("stop") is True or record.get("stopstatus") is True
    return not is_frozen and not is_stopped


def vendor_identity(record: dict[str, Any]) -> str:
    return as_text(record.get("id")) or f"{as_text(record.get('code'))}:{as_text(record.get('name'))}"


def vendor_preference_score(record: dict[str, Any], org_id: str) -> tuple[int, int, str]:
    record_org = as_text(record.get("orgId") or record.get("org"))
    return (
        1 if org_id and record_org == org_id else 0,
        1 if record_org == "666666" else 0,
        as_text(record.get("pubts")),
    )


def dedupe_vendors(records: list[dict[str, Any]], org_id: str) -> list[dict[str, Any]]:
    unique: dict[str, dict[str, Any]] = {}
    for record in records:
        key = vendor_identity(record)
        if not key:
            continue
        if key not in unique or vendor_preference_score(record, org_id) > vendor_preference_score(unique[key], org_id):
            unique[key] = record
    return list(unique.values())


def choose_bank(banks: Any) -> dict[str, Any]:
    rows = [bank for bank in banks if isinstance(bank, dict)] if isinstance(banks, list) else []
    active = [bank for bank in rows if bank.get("stopstatus") is not True]
    for bank in active:
        if bank.get("defaultbank") is True:
            return bank
    return active[0] if active else {}


def openaccountbank_name(bank: dict[str, Any]) -> str:
    value = as_text(bank.get("openaccountbank_name"))
    if value:
        return value
    nested = bank.get("openaccountbank")
    return as_text(nested) if isinstance(nested, dict) else ""


def vendor_cache_row(record: dict[str, Any]) -> dict[str, Any]:
    bank = choose_bank(record.get("vendorbanks"))
    return {
        "id": as_text(record.get("id")),
        "code": as_text(record.get("code")),
        "name": as_text(record.get("name")),
        "creditcode": as_text(record.get("creditcode")),
        "address": as_text(record.get("address") or record.get("vendoraddress")),
        "contactphone": as_text(record.get("contactphone") or record.get("vendorphone")),
        "openaccountbankName": openaccountbank_name(bank),
        "bankAccount": as_text(bank.get("account")),
        "bankAccountName": as_text(bank.get("accountname")),
        "vendorFax": as_text(record.get("vendorfax")),
        "org": as_text(record.get("orgId") or record.get("org")),
        "orgName": as_text(record.get("org_name") or record.get("orgName")),
        "accessstatus": as_text(record.get("accessstatus")),
        "freezestatus": record.get("freezestatus"),
        "pubts": as_text(record.get("pubts")),
    }


def supplier_row_from_render_data(render_data: dict[str, Any]) -> dict[str, Any]:
    name = as_text(render_data.get("supplierName"))
    if not name:
        return {}
    return {
        "id": "",
        "code": "",
        "name": name,
        "creditcode": as_text(render_data.get("supplierTaxNo")),
        "address": as_text(render_data.get("supplierAddress")),
        "contactphone": as_text(render_data.get("supplierPhone")),
        "openaccountbankName": as_text(render_data.get("supplierBank")),
        "bankAccount": as_text(render_data.get("supplierAccount")),
        "bankAccountName": "",
        "vendorFax": as_text(render_data.get("supplierFax")),
        "org": "",
        "orgName": "",
        "accessstatus": "",
        "freezestatus": "",
        "pubts": now_shanghai().isoformat(timespec="seconds"),
    }


def _row_key_by_id(row: dict[str, Any]) -> str:
    return as_text(row.get("id"))


def _row_key_by_name(row: dict[str, Any]) -> str:
    return normalize_supplier_name(row.get("name"))


def append_new_supplier_rows(existing_rows: list[dict[str, Any]], incoming_rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], dict[str, int]]:
    merged = [dict(row) for row in existing_rows]
    existing_ids = {_row_key_by_id(row) for row in merged if _row_key_by_id(row)}
    existing_names = {_row_key_by_name(row) for row in merged if not _row_key_by_id(row) and _row_key_by_name(row)}
    stats = {"existingCacheCount": len(merged), "addedVendorCount": 0, "skippedVendorCount": 0}
    for row in incoming_rows:
        candidate = dict(row)
        row_id = _row_key_by_id(candidate)
        name_key = _row_key_by_name(candidate)
        exists = bool(row_id and row_id in existing_ids) or bool(not row_id and name_key and name_key in existing_names)
        if exists:
            stats["skippedVendorCount"] += 1
            continue
        merged.append(candidate)
        stats["addedVendorCount"] += 1
        if row_id:
            existing_ids.add(row_id)
        elif name_key:
            existing_names.add(name_key)
    stats["cacheVendorCount"] = len(merged)
    return merged, stats


def upsert_confirmed_supplier_row(existing_rows: list[dict[str, Any]], supplier_row: dict[str, Any]) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    if not supplier_row or not as_text(supplier_row.get("name")):
        return existing_rows, {"updated": False, "added": False, "reason": "missing_supplier_name"}
    merged = [dict(row) for row in existing_rows]
    supplier_id = _row_key_by_id(supplier_row)
    supplier_name = _row_key_by_name(supplier_row)
    matched_index: int | None = None
    for index, row in enumerate(merged):
        if supplier_id and _row_key_by_id(row) == supplier_id:
            matched_index = index
            break
        if supplier_name and _row_key_by_name(row) == supplier_name:
            matched_index = index
            break
    if matched_index is None:
        merged.append(dict(supplier_row))
        return merged, {"updated": False, "added": True, "supplierName": supplier_row.get("name")}

    changed_fields: list[str] = []
    target = dict(merged[matched_index])
    for key, value in supplier_row.items():
        text = as_text(value)
        if not text:
            continue
        if as_text(target.get(key)) != text:
            target[key] = text
            changed_fields.append(key)
    merged[matched_index] = target
    return merged, {
        "updated": bool(changed_fields),
        "added": False,
        "supplierName": target.get("name"),
        "changedFields": changed_fields,
    }


def supplier_patch_from_cache(extracted: dict[str, Any], cache_rows: list[dict[str, Any]]) -> dict[str, Any]:
    supplier_name = as_text(extracted.get("supplierName"))
    if not supplier_name:
        return {"matched": False, "patch": {}, "reason": "missing_supplier_name"}
    name_key = normalize_supplier_name(supplier_name)
    matches = [row for row in cache_rows if _row_key_by_name(row) == name_key]
    if not matches:
        return {"matched": False, "patch": {}, "reason": "not_found", "supplierName": supplier_name}
    if len(matches) > 1:
        return {"matched": False, "patch": {}, "reason": "ambiguous", "supplierName": supplier_name}

    row = matches[0]
    patch: dict[str, str] = {}
    for field_key, row_key in SUPPLIER_FIELD_MAP.items():
        if not is_blank(extracted.get(field_key)):
            continue
        value = as_text(row.get(row_key))
        if value:
            patch[field_key] = value
    return {
        "matched": True,
        "supplierName": supplier_name,
        "cacheSupplierName": row.get("name"),
        "patch": patch,
        "patchedFields": sorted(patch),
    }


def apply_supplier_patch(extracted: dict[str, Any], patch: dict[str, Any]) -> set[str]:
    changed: set[str] = set()
    values = patch.get("patch") if isinstance(patch.get("patch"), dict) else {}
    for key, value in values.items():
        if is_blank(extracted.get(key)) and not is_blank(value):
            extracted[key] = as_text(value)
            changed.add(key)
    return changed


def now_shanghai() -> datetime:
    return datetime.now(timezone(timedelta(hours=8)))


def write_supplier_cache_xlsx(path: Path, rows: list[dict[str, Any]], manifest: dict[str, Any]) -> None:
    workbook = Workbook()
    vendors_sheet = workbook.active
    vendors_sheet.title = "供应商"
    vendors_sheet.append([label for _key, label in VENDOR_COLUMNS])
    for row in rows:
        vendors_sheet.append([row.get(key, "") for key, _label in VENDOR_COLUMNS])

    manifest_labels = {
        "syncedAt": "同步时间",
        "sourceRecordCount": "用友原始记录数",
        "fetchedRecordCount": "实际抓取记录数",
        "availableRecordCount": "可用记录数",
        "uniqueVendorCount": "去重后供应商数",
        "existingCacheCount": "原缓存供应商数",
        "addedVendorCount": "本次新增供应商数",
        "skippedVendorCount": "本次已存在供应商数",
        "cacheVendorCount": "缓存供应商总数",
        "pageSize": "分页大小",
        "maxPages": "最大抓取页数",
        "sourceApi": "来源接口",
        "tokenExpire": "令牌有效期秒数",
    }
    manifest_sheet = workbook.create_sheet("同步信息")
    manifest_sheet.append(["项目", "值"])
    for key, value in manifest.items():
        label = manifest_labels.get(key, key)
        display_value = json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else value
        manifest_sheet.append([label, display_value])
    workbook.save(path)


def read_supplier_cache_xlsx(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    if "供应商" not in workbook.sheetnames:
        return []
    sheet = workbook["供应商"]
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return []
    keys = [VENDOR_COLUMN_LABEL_TO_KEY.get(as_text(label), as_text(label)) for label in header]
    rows: list[dict[str, Any]] = []
    for values in rows_iter:
        row = {key: as_text(value) for key, value in zip(keys, values) if key}
        if as_text(row.get("id")) or as_text(row.get("name")):
            rows.append({key: row.get(key, "") for key, _label in VENDOR_COLUMNS})
    return rows


def sync_suppliers_to_xlsx(output_dir: Path, existing_rows: list[dict[str, Any]] | None = None) -> dict[str, Any]:
    gateway_url, token_url = resolve_endpoints()
    access_token, expire = get_access_token(token_url)
    page_size = env_int("YONBIP_VENDOR_PAGE_SIZE", DEFAULT_PAGE_SIZE)
    org_id = optional_env("YONBIP_ORG_ID")

    all_records: list[dict[str, Any]] = []
    source_record_count = 0
    page_index = 1
    while True:
        page = query_vendor_page(gateway_url, access_token, page_index, page_size)
        if page_index == 1:
            source_record_count = page["recordCount"]
        all_records.extend(page["records"])
        page_count = page["pageCount"]
        if not page["records"] or (page_count and page_index >= page_count):
            break
        if source_record_count and len(all_records) >= source_record_count:
            break
        page_index += 1

    available_records = [record for record in all_records if vendor_is_available(record)]
    unique_records = dedupe_vendors(available_records, org_id)
    incoming_rows = [vendor_cache_row(record) for record in unique_records]
    rows, merge_stats = append_new_supplier_rows(existing_rows or [], incoming_rows)
    synced_at = now_shanghai().isoformat(timespec="seconds")
    file_name = SUPPLIER_CACHE_FILE_NAME
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / file_name
    manifest = {
        "syncedAt": synced_at,
        "sourceRecordCount": source_record_count or len(all_records),
        "fetchedRecordCount": len(all_records),
        "availableRecordCount": len(available_records),
        "uniqueVendorCount": len(incoming_rows),
        **merge_stats,
        "pageSize": page_size,
        "sourceApi": "vendor/queryByPage",
        "tokenExpire": expire,
    }
    write_supplier_cache_xlsx(output_path, rows, manifest)
    return {
        "fileName": file_name,
        "path": output_path,
        **manifest,
    }
